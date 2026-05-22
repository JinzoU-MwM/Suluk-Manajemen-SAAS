"""
Excel Service - Handles Excel template manipulation and data export
"""
import logging
from pathlib import Path
from typing import List, Dict
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from ..schemas import DocumentData
from ..utils.helpers import convert_date_format
from ..config import DEFAULT_NATIONALITY, DEFAULT_IDENTITY_TYPE_KTP, DEFAULT_IDENTITY_TYPE_PASSPORT

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for handling Excel operations"""
    
    # Exact column headers for Siskopatuh system
    COLUMN_HEADERS = [
        "No Identitas",
        "Nama (Sesuai Dengan nama Pada Kartu Vaksin)",
        "Tempat Lahir",
        "Tanggal Lahir(yyyy-mm-dd)",
        "Alamat",
        "Provinsi",
        "Kabupaten",
        "Kecamatan",
        "Kelurahan",
        "No Paspor",
        "Tanggal Dikeluarkan Paspor(yyyy-mm-dd)",
        "Kota Paspor",
        "No Visa",
        "Jenis Identitas",
        "KewargaNegaraan"
    ]
    
    def __init__(self):
        """Initialize Excel service"""
        self.workbook = None
        self.worksheet = None
    
    def load_template(self, template_path: str) -> None:
        """Load Excel template (.xlsm)"""
        logger.info(f"Loading template from: {template_path}")
        try:
            self.workbook = load_workbook(template_path, keep_vba=True)
            if self.workbook.sheetnames:
                self.worksheet = self.workbook.active
            else:
                self.worksheet = self.workbook.create_sheet("Data Jamaah")
            logger.info(f"Template loaded successfully. Active sheet: {self.worksheet.title}")
        except Exception as e:
            logger.error(f"Error loading template: {e}")
            raise
    
    def create_new_workbook(self) -> None:
        """Create a new Excel workbook with headers"""
        logger.info("Creating new workbook")
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Data Jamaah"
        
        for col_idx, header in enumerate(self.COLUMN_HEADERS, start=1):
            self.worksheet.cell(row=1, column=col_idx, value=header)
    
    def map_to_excel_row(self, data: DocumentData) -> Dict[str, str]:
        """Map DocumentData to Excel row format"""
        row = {}
        
        # Determine identity type
        if data.document_type == "KTP":
            jenis_identitas = DEFAULT_IDENTITY_TYPE_KTP
            no_identitas = data.nik
        elif data.document_type == "PASSPORT":
            jenis_identitas = DEFAULT_IDENTITY_TYPE_PASSPORT
            no_identitas = data.passport_number
        else:
            jenis_identitas = DEFAULT_IDENTITY_TYPE_PASSPORT
            no_identitas = data.passport_number or data.nik
        
        # Map fields
        row["No Identitas"] = no_identitas
        row["Nama (Sesuai Dengan nama Pada Kartu Vaksin)"] = data.name
        row["Tempat Lahir"] = data.place_of_birth
        
        # Date format check
        if data.date_of_birth:
            if '-' in data.date_of_birth:
                parts = data.date_of_birth.split('-')
                if len(parts[0]) == 2:  # DD-MM-YYYY
                    row["Tanggal Lahir(yyyy-mm-dd)"] = convert_date_format(
                        data.date_of_birth, "%d-%m-%Y", "%Y-%m-%d"
                    )
                else:
                    row["Tanggal Lahir(yyyy-mm-dd)"] = data.date_of_birth
            else:
                row["Tanggal Lahir(yyyy-mm-dd)"] = data.date_of_birth
        else:
            row["Tanggal Lahir(yyyy-mm-dd)"] = None
        
        row["Alamat"] = data.address
        row["Provinsi"] = data.provinsi
        row["Kabupaten"] = data.kabupaten
        row["Kecamatan"] = data.kecamatan
        row["Kelurahan"] = data.kelurahan
        row["No Paspor"] = data.passport_number
        row["Tanggal Dikeluarkan Paspor(yyyy-mm-dd)"] = data.date_of_issue
        row["Kota Paspor"] = data.city_of_issue
        row["No Visa"] = data.visa_number
        row["Jenis Identitas"] = jenis_identitas
        row["KewargaNegaraan"] = DEFAULT_NATIONALITY
        
        return row
    
    def create_dataframe(self, parsed_data: List[DocumentData]) -> pd.DataFrame:
        """Create pandas DataFrame from parsed data"""
        logger.info(f"Creating DataFrame from {len(parsed_data)} documents")
        rows = []
        for data in parsed_data:
            row = self.map_to_excel_row(data)
            rows.append(row)
        df = pd.DataFrame(rows, columns=self.COLUMN_HEADERS)
        return df
    
    def append_to_template(self, dataframe: pd.DataFrame) -> None:
        """Append DataFrame data to the loaded template"""
        if self.worksheet is None:
            raise ValueError("No worksheet loaded.")
        
        next_row = self.worksheet.max_row + 1
        if self.worksheet.max_row == 1 and self.worksheet.cell(1, 1).value is None:
            for col_idx, header in enumerate(self.COLUMN_HEADERS, start=1):
                self.worksheet.cell(row=1, column=col_idx, value=header)
            next_row = 2
        
        for row_idx, row_data in dataframe.iterrows():
            for col_idx, header in enumerate(self.COLUMN_HEADERS, start=1):
                value = row_data[header]
                self.worksheet.cell(row=next_row, column=col_idx, value=value)
            next_row += 1
    
    def save_xlsm(self, output_path: str) -> None:
        """Save workbook as .xlsm file"""
        if self.workbook is None:
            raise ValueError("No workbook to save.")
        
        output_path = Path(output_path)
        if output_path.suffix != '.xlsm':
            output_path = output_path.with_suffix('.xlsm')
        
        self.workbook.save(output_path)
        logger.info(f"Workbook saved to: {output_path}")
