"""
Export Template Router - Custom Excel template upload and export.
"""
import os
import uuid
from pathlib import Path
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.auth import get_current_user
from app.models.user import User
from app.models.export_template import ExportTemplate
from app.models.group import Group, GroupMember
from app.config import UPLOAD_DIR, MAX_FILE_SIZE

router = APIRouter(prefix="/export-templates", tags=["export"])
TEMPLATE_UPLOAD_DIR = (UPLOAD_DIR / "templates").resolve()


def _ensure_template_upload_dir() -> None:
    """Ensure template upload directory exists when upload endpoint is called."""
    try:
        TEMPLATE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    except PermissionError as exc:
        raise HTTPException(
            status_code=503,
            detail="Template storage is not writable on this server",
        ) from exc

# Header mapping for auto-detection
HEADER_MAPPINGS = {
    # Indonesian
    "nama": "nama",
    "nama lengkap": "nama",
    "nama jamaah": "nama",
    "no. identitas": "no_identitas",
    "nik": "no_identitas",
    "no ktp": "no_identitas",
    "no paspor": "no_paspor",
    "nomor paspor": "no_paspor",
    "passport no": "no_paspor",
    "no passport": "no_paspor",
    "tanggal lahir": "tanggal_lahir",
    "tgl lahir": "tanggal_lahir",
    "dob": "tanggal_lahir",
    "date of birth": "tanggal_lahir",
    "tempat lahir": "tempat_lahir",
    "alamat": "alamat",
    "alamat lengkap": "alamat",
    "provinsi": "provinsi",
    "kabupaten": "kabupaten",
    "kota": "kabupaten",
    "kecamatan": "kecamatan",
    "kelurahan": "kelurahan",
    "desa": "kelurahan",
    "no telepon": "no_telepon",
    "no hp": "no_hp",
    "phone": "no_hp",
    "telepon": "no_telepon",
    "jenis kelamin": "jenis_kelamin",
    "gender": "jenis_kelamin",
    "jk": "jenis_kelamin",
    "kewarganegaraan": "kewarganegaraan",
    "status pernikahan": "status_pernikahan",
    "status": "status_pernikahan",
    "pekerjaan": "pekerjaan",
    "pendidikan": "pendidikan",
    "visa": "no_visa",
    "no visa": "no_visa",
    "tanggal visa": "tanggal_visa",
    "asuransi": "asuransi",
    "no polis": "no_polis",
}


class ConfigureTemplateRequest(BaseModel):
    name: str | None = None
    column_mapping: dict
    header_row: int = 1
    data_start_row: int = 2


def _resolve_template_path(path_value: str) -> Path:
    """Resolve and enforce template path boundary inside TEMPLATE_UPLOAD_DIR."""
    resolved = Path(path_value).resolve()
    if TEMPLATE_UPLOAD_DIR not in resolved.parents:
        raise HTTPException(400, "Invalid template path")
    return resolved


@router.post("/upload")
async def upload_template(
    file: UploadFile = File(...),
    name: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload Excel template and return server-side template handle."""
    _ensure_template_upload_dir()

    original_name = Path(file.filename or "").name
    ext = Path(original_name).suffix.lower()
    if ext not in {".xlsx", ".xls", ".xlsm"}:
        raise HTTPException(400, "File must be Excel (.xlsx, .xls, .xlsm)")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "Template file too large")

    safe_filename = f"{current_user.id}_{uuid.uuid4().hex}{ext}"
    file_path = (TEMPLATE_UPLOAD_DIR / safe_filename).resolve()
    with open(file_path, "wb") as f:
        f.write(content)

    # Load workbook and detect headers
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        header_row = 1
        headers = []
        for cell in ws[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Auto-detect mapping
        mapping = {}
        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()
            if header_lower in HEADER_MAPPINGS:
                mapping[idx] = HEADER_MAPPINGS[header_lower]

        template = ExportTemplate(
            user_id=current_user.id,
            name=(name or "Template").strip(),
            file_path=str(file_path),
            column_mapping=mapping,
            header_row=1,
            data_start_row=2,
        )
        db.add(template)
        db.commit()
        db.refresh(template)

        return {
            "template_id": template.id,
            "headers": headers,
            "auto_mapping": mapping,
            "detected_count": len(mapping),
            "total_columns": len(headers),
            "name": template.name,
        }
    except Exception as e:
        if file_path.exists():
            os.remove(file_path)
        raise HTTPException(400, f"Failed to read Excel file: {str(e)}")


@router.post("/{template_id}/configure")
def configure_template(
    template_id: int,
    body: ConfigureTemplateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Configure an uploaded template by id (server-issued handle)."""
    template = (
        db.query(ExportTemplate)
        .filter(
            ExportTemplate.id == template_id,
            ExportTemplate.user_id == current_user.id,
        )
        .first()
    )
    if not template:
        raise HTTPException(404, "Template not found")

    template_path = _resolve_template_path(template.file_path)
    if not template_path.exists():
        raise HTTPException(400, "Template file not found")

    template.name = (body.name or template.name).strip()
    template.column_mapping = body.column_mapping
    template.header_row = body.header_row
    template.data_start_row = body.data_start_row
    db.commit()
    db.refresh(template)

    return {
        "id": template.id,
        "name": template.name,
        "column_count": len(body.column_mapping),
    }


@router.get("/")
def list_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List saved templates."""
    templates = (
        db.query(ExportTemplate)
        .filter(ExportTemplate.user_id == current_user.id)
        .order_by(ExportTemplate.created_at.desc())
        .all()
    )

    return {
        "templates": [
            {
                "id": t.id,
                "name": t.name,
                "created_at": t.created_at.isoformat(),
            }
            for t in templates
        ]
    }


@router.get("/{template_id}")
def get_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get template details."""
    template = (
        db.query(ExportTemplate)
        .filter(
            ExportTemplate.id == template_id,
            ExportTemplate.user_id == current_user.id,
        )
        .first()
    )

    if not template:
        raise HTTPException(404, "Template not found")

    return {
        "id": template.id,
        "name": template.name,
        "column_mapping": template.column_mapping,
        "header_row": template.header_row,
        "data_start_row": template.data_start_row,
    }


@router.delete("/{template_id}")
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a template."""
    template = (
        db.query(ExportTemplate)
        .filter(
            ExportTemplate.id == template_id,
            ExportTemplate.user_id == current_user.id,
        )
        .first()
    )

    if not template:
        raise HTTPException(404, "Template not found")

    template_path = _resolve_template_path(template.file_path)
    if template_path.exists():
        os.remove(template_path)

    db.delete(template)
    db.commit()

    return {"success": True}


@router.post("/export/{group_id}")
def export_with_template(
    group_id: int,
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export group data using a saved template."""
    # Verify group ownership
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group or group.user_id != current_user.id:
        raise HTTPException(403, "Not authorized")

    # Get template
    template = (
        db.query(ExportTemplate)
        .filter(
            ExportTemplate.id == template_id,
            ExportTemplate.user_id == current_user.id,
        )
        .first()
    )
    if not template:
        raise HTTPException(404, "Template not found")

    # Get members
    members = (
        db.query(GroupMember)
        .filter(GroupMember.group_id == group_id)
        .order_by(GroupMember.nama)
        .all()
    )

    # Load template
    template_path = _resolve_template_path(template.file_path)
    if not template_path.exists():
        raise HTTPException(400, "Template file not found")
    wb = load_workbook(template_path)
    ws = wb.active

    # Write data
    row_idx = template.data_start_row
    for member in members:
        for col_idx, field_name in template.column_mapping.items():
            col_num = int(col_idx)
            col_letter = get_column_letter(col_num + 1)  # 1-indexed
            value = getattr(member, field_name, "")
            ws[f"{col_letter}{row_idx}"] = value or ""
        row_idx += 1

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{group.name.replace(' ', '_')}_export.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
