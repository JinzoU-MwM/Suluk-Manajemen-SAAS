"""
Excel Router — /generate-excel/ and /download/ endpoints.
Handles Excel file generation from verified data and file downloads.
"""
import io
import logging
from datetime import datetime

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.schemas import GenerateExcelRequest
from app.config import OUTPUT_DIR
from app.database import get_db
from app.auth import get_current_user, check_access
from app.models.user import User
from app.services.siskopatuh_validation import (
    normalize_items_to_siskopatuh_dropdowns,
    validate_items_against_siskopatuh_dropdowns,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Excel"])

# 32-column headers matching Siskopatuh Excel structure
EXCEL_HEADERS = [
    "Title",
    "Nama (Sesuai Dengan nama Pada Kartu Vaksin)",
    "Nama Ayah",
    "Jenis Identitas",
    "No Identitas",
    "Nama Paspor",
    "No Paspor",
    "Tanggal Dikeluarkan Paspor(yyyy-mm-dd)",
    "Kota Paspor",
    "Tempat Lahir",
    "Tanggal Lahir(yyyy-mm-dd)",
    "Alamat",
    "Provinsi",
    "Kabupaten",
    "Kecamatan",
    "Kelurahan",
    "No. Telepon",
    "No Hp",
    "KewargaNegaraan",
    "Status Pernikahan",
    "Pendidikan",
    "Pekerjaan",
    "Provider Visa",
    "No Visa",
    "Tanggal Berlaku Visa (yyyy-mm-dd)",
    "Tanggal Akhir  Visa (yyyy-mm-dd)",
    "Asuransi",
    "No Polis",
    "Tanggal Input Polis (yyyy-mm-dd)",
    "Tanggal Awal Polis (yyyy-mm-dd)",
    "Tanggal Akhir Polis (yyyy-mm-dd)",
    "No BPJS",
]


def _item_to_row(item) -> list:
    """Convert an ExtractedDataItem to a list of values matching EXCEL_HEADERS order."""
    return [
        item.title, item.nama, item.nama_ayah, item.jenis_identitas,
        item.no_identitas, item.nama_paspor, item.no_paspor, item.tanggal_paspor,
        item.kota_paspor, item.tempat_lahir, item.tanggal_lahir, item.alamat,
        item.provinsi, item.kabupaten, item.kecamatan, item.kelurahan,
        item.no_telepon, item.no_hp, item.kewarganegaraan, item.status_pernikahan,
        item.pendidikan, item.pekerjaan, item.provider_visa, item.no_visa,
        item.tanggal_visa, item.tanggal_visa_akhir, item.asuransi, item.no_polis,
        item.tanggal_input_polis, item.tanggal_awal_polis, item.tanggal_akhir_polis,
        item.no_bpjs,
    ]


@router.post("/generate-excel/")
async def generate_excel(
    request: GenerateExcelRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate Excel file from verified/edited data (32 columns). Pro only."""
    # Check if user has Pro plan
    access = check_access(db, user)
    if access["plan"] != "pro":
        raise HTTPException(
            status_code=403,
            detail="Fitur export Excel hanya tersedia untuk pengguna Pro. Upgrade untuk mengakses fitur ini.",
        )

    try:
        logger.info(f"Generating Excel for {len(request.data)} rows")

        if not request.data:
            raise HTTPException(status_code=400, detail="No data provided")

        normalize_items_to_siskopatuh_dropdowns(request.data)
        dropdown_errors = validate_items_against_siskopatuh_dropdowns(request.data)
        if dropdown_errors:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Data tidak sesuai opsi dropdown template Siskopatuh.",
                    "errors": dropdown_errors[:50],
                    "total_errors": len(dropdown_errors),
                },
            )

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Jamaah"

        # Write headers
        for col, header in enumerate(EXCEL_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data rows
        for row_idx, item in enumerate(request.data, 2):
            row_values = _item_to_row(item)
            for col, value in enumerate(row_values, 1):
                ws.cell(row=row_idx, column=col, value=value)

        output_filename = f"jamaah_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={output_filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating Excel file: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error generating Excel file: {str(e)}"
        )


@router.get("/download/{filename}")
async def download_file(filename: str):
    """Download a previously generated file."""
    file_path = OUTPUT_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
