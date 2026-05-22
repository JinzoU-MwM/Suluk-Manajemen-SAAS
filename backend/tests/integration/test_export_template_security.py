from datetime import timedelta
from pathlib import Path
from io import BytesIO

from app.auth import create_access_token, hash_password
from app.models.export_template import ExportTemplate
from app.models.user import User, Subscription, PlanType, SubscriptionStatus, utc_now
from openpyxl import Workbook


def _create_user(db_session, email: str) -> User:
    user = User(
        email=email,
        name="Template User",
        password_hash=hash_password("password123"),
        is_active=True,
        email_verified=True,
    )
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)

    db_session.add(
        Subscription(
            user_id=user.id,
            plan=PlanType.FREE,
            status=SubscriptionStatus.TRIAL,
            trial_start=utc_now(),
            trial_end=utc_now() + timedelta(days=7),
        )
    )
    db_session.commit()
    return user


def _auth_headers(user: User) -> dict:
    token = create_access_token(data={"sub": str(user.id), "email": user.email})
    return {"Authorization": f"Bearer {token}"}


def test_upload_template_returns_server_handle_not_file_path(client, db_session):
    user = _create_user(db_session, "exporter@example.com")
    headers = _auth_headers(user)

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Nama"
    ws["B1"] = "No Paspor"
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    response = client.post(
        "/export-templates/upload",
        headers=headers,
        files={"file": ("template.xlsx", buff.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"name": "Sample Template"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert "template_id" in payload
    assert "file_path" not in payload


def test_delete_template_rejects_path_outside_template_root(client, db_session):
    user = _create_user(db_session, "pathguard@example.com")
    headers = _auth_headers(user)

    external_path = str((Path(__file__).resolve().parent / "outside.xlsx").resolve())
    template = ExportTemplate(
        user_id=user.id,
        name="Unsafe",
        file_path=external_path,
        column_mapping={},
    )
    db_session.add(template)
    db_session.commit()
    db_session.refresh(template)

    response = client.delete(f"/export-templates/{template.id}", headers=headers)
    assert response.status_code == 400
    assert "Invalid template path" in response.text
